"""Модуль для записей списков и словарей в файл."""

import csv
import datetime
import shutil
from abc import ABC, abstractmethod
from dataclasses import dataclass
from enum import Enum, unique
from os import path, remove
from pathlib import PosixPath
from typing import NewType, Type, Union

from django.conf import settings
from django.template import Context
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .document_generator import DocumentData, DocumentGenerator

FormatValue = NewType('FormatValue', Union[str, datetime.datetime, datetime.date, float, int])
FormatType = NewType('FormatType', Type[FormatValue])


@dataclass
class HeaderOption:
    """Опция для форматирования запроса."""

    name: str
    format: FormatType


@dataclass
class HeaderOptions:
    """Набор опций заголовков."""

    code: str
    option: HeaderOption


class BaseWriter(ABC):
    """Интерфейс, описывающий контракт для записи информации в файл."""

    def __init__(self, path_file: PosixPath) -> None:
        """Интерфейс инициализации."""
        self.path_file = path_file

    @property
    def relpath(self) -> str:
        """Относительный путь."""
        return path.relpath(self.path_file, settings.BASE_DIR)

    @abstractmethod
    def write_from_dict(self, data: list[dict], fieldnames: list[str]) -> None:
        """Записывающий объект представляет собой.

        - data - [
            {'p1': 1, 'p2': 2, 'p3': 3, 'p4': 4},
            ...
        ]
        - fieldnames - ['p1', 'p2', 'p3', 'p4']
        """
        ...


class CsvWriter(BaseWriter):
    """Записываем в csv файл."""

    def __init__(self, path_file: PosixPath) -> None:
        """Инициализация записи в csv."""
        super().__init__(path_file)

    def write_from_dict(self, data: list[dict], fieldnames: list[str]) -> None:
        """Записываем из словаря."""
        with open(self.path_file, 'w') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=';')
            writer.writeheader()
            writer.writerows(data)


class ExcelWriter(BaseWriter):
    """Записываем xlsx файл."""

    def __init__(self, path_file: PosixPath) -> None:
        """Инициализация записи в excel."""
        super().__init__(path_file)

    def write_from_dict(self, data: list[dict], fieldnames: list[str]) -> None:
        """Записываем из словаря."""
        workbook: Workbook = Workbook()
        workbook.remove(workbook.active)
        work_sheet: Worksheet = workbook.create_sheet('Выгрузка')
        for column_index, header in enumerate(fieldnames, 1):
            work_sheet.cell(1, column_index, header)
        row_index: int = 2
        for row in data:
            for column_index, header in enumerate(fieldnames, 1):
                work_sheet.cell(row_index, column_index, row[header])
            row_index += 1
        workbook.save(self.path_file)


class PdfWriter(BaseWriter):
    """Записываем в pdf файл."""

    document: PosixPath = settings.BASE_DIR / 'apps' / 'coins' / 'templates' / 'document.docx'
    template: PosixPath = settings.BASE_DIR / 'apps' / 'coins' / 'templates' / 'document.xml'

    def __init__(self, path_file: PosixPath) -> None:
        """Инициализация записи в pdf."""
        super().__init__(path_file)

    def write_from_dict(self, data: list[dict], fieldnames: list[str]) -> None:
        """Записываем из словаря."""
        context: Context = Context({
            'headers': fieldnames,
            'data': data
        })
        document_generator: DocumentGenerator = DocumentGenerator(context, self.template, self.document)
        document: DocumentData = document_generator.generate_pdf(True)
        if path.exists(document.full_path):
            shutil.copy(document.full_path, self.path_file)
            remove(document.full_path)


@unique
class Writers(Enum):
    """Тип перечисления доступных писателей."""

    CSV: str = 'csv'
    EXCEL: str = 'xlsx'
    PDF: str = 'pdf'


class ContainerWriter:
    """Фабрика для записи файлов."""

    available_writers: dict[Writers, BaseWriter] = {
        Writers.CSV: CsvWriter,
        Writers.EXCEL: ExcelWriter,
        Writers.PDF: PdfWriter,
    }

    def __init__(self, headers: list[HeaderOptions], path_file: PosixPath, writer_type: Writers = Writers.CSV) -> None:
        """Инициализация записи в файл."""
        self.headers: list[HeaderOptions] = headers
        self.writer: BaseWriter = self.available_writers.get(writer_type, CsvWriter)(path_file)

    @property
    def relpath(self) -> str:
        """Проброс относительного пути."""
        return self.writer.relpath

    def write_dict(self, raw_data: list[dict]) -> str:
        """Записываем информацию в файл из словаря."""
        data, fieldnames = self._format_data_dict(raw_data)
        self.writer.write_from_dict(data, fieldnames)
        return self.writer.relpath

    def _format_data_dict(self, raw_data: list[dict]) -> tuple[list[dict], list[str]]:
        """Обеспечивает форматирование и последовательность данных."""
        return [
            {
                header.option.name: self._format_value(row.get(header.code), header.option.format)
                for header in self.headers
            } for row in raw_data
        ], [header.option.name for header in self.headers],

    @staticmethod
    def _format_value(value: FormatValue | None, format_type: FormatType) -> str:
        """Обеспечиваем форматирование в зависимости от типа. Форматирование должно быть единое для всего сайта."""
        if value is None:
            return ''
        match format_type:
            case datetime.date:
                return value.strftime('%d.%m.%Y')
            case _:
                return value


__all__: list[str] = [
    'Writers',
    'ContainerWriter',
    'HeaderOptions',
    'HeaderOption',
]
